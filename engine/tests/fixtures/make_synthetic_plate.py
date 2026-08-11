"""Generate the synthetic SRB plate fixture (synthetic_srb_plate.xlsx).

Fully fabricated data with analytically known ground truth; no
measurements from any real experiment. Two fictional cell lines follow
exact 4PL dose-response curves (hill = 1):

    corrected(d) = BOTTOM + (TOP - BOTTOM) / (1 + d / ic50)

Replicate wells get offsets (-0.003, 0, +0.003) that cancel in the mean,
so per-dose corrected means equal the 4PL values exactly. Well value =
corrected + BLANK. Run this script to (re)create the fixture and print
the ground-truth constants used by test_plate.py.
"""

from pathlib import Path

from openpyxl import Workbook

BLANK = 0.100
TOP, BOTTOM = 2.4, 0.06
DOSES = [0.0, 5.0, 3.0, 2.0, 1.0, 0.7, 0.5, 0.3, 0.1, 0.05]  # cols 2..11
GROUPS = [("Line S", ["B", "C", "D"], 0.5), ("Line R", ["E", "F", "G"], 2.0)]
OFFSETS = [-0.003, 0.0, 0.003]


def corrected(dose: float, ic50: float) -> float:
    if dose == 0.0:
        return TOP
    return BOTTOM + (TOP - BOTTOM) / (1 + dose / ic50)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "plate"
    # Column labels 1..12 in row 1, row labels A..H in column A
    for c in range(1, 13):
        ws.cell(row=1, column=c + 1, value=c)
    for r, letter in enumerate("ABCDEFGH", start=2):
        ws.cell(row=r, column=1, value=letter)
    # Filler wells (edges) sit near the blank level
    for r in range(2, 10):
        for c in range(2, 14):
            ws.cell(row=r, column=c, value=0.095)
    # Blank wells H3..H5
    for c in (3, 4, 5):
        ws.cell(row=9, column=c + 1, value=BLANK)
    # Dose wells
    for name, rows, ic50 in GROUPS:
        for ri, row_letter in enumerate(rows):
            r = 2 + "ABCDEFGH".index(row_letter)
            for ci, dose in enumerate(DOSES):
                v = corrected(dose, ic50) + BLANK + OFFSETS[ri]
                ws.cell(row=r, column=ci + 3, value=round(v, 6))

    out = Path(__file__).parent / "synthetic_srb_plate.xlsx"
    wb.save(out)
    print(f"wrote {out}")
    for name, _rows, ic50 in GROUPS:
        truth = {d: round(corrected(d, ic50), 7) for d in DOSES if d != 0.0}
        print(name, "ic50", ic50, "corrected means:", truth)
        print(name, "control mean:", corrected(0.0, ic50))


if __name__ == "__main__":
    main()
